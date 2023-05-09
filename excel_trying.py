import csv
import os
import time

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle


def what_if_its_open(xlsx_file_name):
    tmp_part = f'_{int(time.time() * 100) % 1000000}'
    try:
        os.rename(xlsx_file_name, xlsx_file_name)
    except OSError:
        index = xlsx_file_name.find('.xlsx')
        final_name = xlsx_file_name[:index] + tmp_part + xlsx_file_name[index:]
        return final_name
    return xlsx_file_name

def create_excel(input_csv_file, output_xlsx_file):
    final_xlsx_name = what_if_its_open(output_xlsx_file)
    print(f' {final_xlsx_name} ', end='')
    with open(input_csv_file, mode="r", encoding="WINDOWS-1251") as csv_file:
        book = openpyxl.Workbook()

        #удаляем дефолтный лист, создаем новый
        book.remove(book.active)
        sheet_1 = book.create_sheet("Ликвидашки")

        fill_1 = PatternFill("solid", fgColor="D8E0BE")
        # Обрабатываем заголовки столбцов
        csv_reader = csv.DictReader(csv_file, delimiter="\t")
        i = 1
        for head_name in csv_reader.fieldnames:
            sheet_1.cell(row=1, column=i).value = head_name
            sheet_1.cell(row=1, column=i).font = Font(b=True)
            sheet_1.cell(row=1, column=i).fill = fill_1
            i += 1
        # применяем автофильтр ко всем столбцам (например от A:G)
        sheet_1.auto_filter.ref = f'A:{"ABCDEFGHIJKLMNOPQRSTUVWXYZ"[i-2] if i<27 else "Z"}'
        sheet_1.column_dimensions["A"].width = 12
        sheet_1.column_dimensions["B"].width = 21
        sheet_1.column_dimensions["C"].width = 8
        sheet_1.column_dimensions["D"].width = 60
        sheet_1.column_dimensions["E"].width = 60
        sheet_1.column_dimensions["F"].width = 19
        sheet_1.column_dimensions["G"].width = 75

        # Читаем данные из cvs и пишем их в лист {sheet_1}
        row_counter = 2
        a_values = []  # : List - список всех значений в строке csv

        my_style_1 = NamedStyle(name="colored_data_style")
        my_style_1.fill = PatternFill("solid", fgColor="D8E0BE")
        my_style_1.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')

        my_style_2 = NamedStyle(name="white_data_style")
        my_style_2.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')

        for row in csv_reader:
            for key, value in row.items():  # read
                a_values.append(value)
            for j in range(0, len(a_values)):  # write
                sheet_1.cell(row=row_counter, column=j+1).value = a_values[j]
                if row_counter % 2 == 1:
                    sheet_1.cell(row=row_counter, column=j + 1).style = my_style_1
                else:
                    sheet_1.cell(row=row_counter, column=j + 1).style = my_style_2
            a_values.clear()
            row_counter += 1

        # Save {book} to file
        book.save(final_xlsx_name)
        book.close()
        os.system("START EXCEL " + final_xlsx_name)


def test_join():
    delimiter = '\t'
    fieldnames = ['ИНН', 'Дата прекращения', 'Причина прекращения', 'Наименование', 'Основной ОКВЭД',
                              'Расшифровка ОКВЭД']

    print(f"{delimiter.join(fieldnames)}\n")


if __name__ == "__main__":
    files_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "output_files")
    csv_file = os.path.join(files_path, "output.csv")
    xlsx_file = os.path.join(files_path, "test.xlsx")

    create_excel(csv_file, xlsx_file)
    # test_join()
